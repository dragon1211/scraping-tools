# !/usr/bin/python
import os
from time import sleep
import requests
import datetime
from random import randint
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook, Workbook
import math


class SolarSystemScript:
    def __init__(self):
        self.url = 'https://www.enfsolar.com/directory/installer/United%20States?page='
        self.start_page = 54
        self.end_page = 71
        self.browser = ''
        self.excel = "Output Data.xlsx"
        self.sheet_name = 'Data Sheet'
        self.remove_file()
        self.create_sheet()
        self.index = 2
        self.header_list = ['URL', 'Name', 'Address', 'Phone', 'Website', 'Country', 'On_Off_Grid', 'Installation Size',
                            'Other Services', 'Operating Area', 'Panel Suppliers', 'Last Update', 'Sales Contracts',
                            'Expansions', 'Financial News', 'Other News']

        self.data_list = []
        self.write_header_table(1, self.header_list)

    def remove_file(self):
        if os.path.exists(self.excel):
            os.remove(self.excel)

    def start_browser(self):
        self.browser = webdriver.Chrome()

    def create_sheet(self):
        wb = Workbook(write_only=True)
        wb.create_sheet(self.sheet_name)
        wb.save(self.excel)

    def write_header_table(self, col_start, header_data):
        wb = load_workbook(self.excel)
        ws = wb.active
        for col_num in range(1, len(header_data) + 1):
            ws.cell(row=1, column=col_num + col_start - 1).value = header_data[col_num - 1]
        wb.save(self.excel)

    def write_to_table(self, row_num, col_start, data):
        wb = load_workbook(self.excel)
        ws = wb.active
        for col_num in range(1, len(data) + 1):
            ws.cell(row=row_num, column=col_num + col_start - 1).value = data[col_num - 1]
        wb.save(self.excel)

    def main(self):
        try:
            print('Start Work')
            self.start_browser()
            for index in range(self.start_page, self.end_page + 1):
                self.browser.get(self.url + str(index))
                urls = []
                table_url = self.browser.find_elements_by_css_selector('body > div.container > div > div > div.mk-body > div.mk-section.clearfix > table > tbody > tr > td:nth-child(1) > a')
                for solar_url in table_url:
                    url = solar_url.get_attribute('href')
                    urls.append(url)
                # print("urls : ", urls)
                for url in urls:
                    self.get_url_data(url)
            # self.get_url_data([])
            print("Work Done ... (*-^)")

        except:
            pass

    def get_url_data(self, url):
        try:
            # # url = 'https://www.enfsolar.com/174-power-global-corporation?directory=installer&utm_source=ENF&utm_medium=United%20States&utm_content=126696&utm_campaign=profiles_installer'
            # url = 'https://www.enfsolar.com/8minute-solar-energy?directory=installer&utm_source=ENF&utm_medium=United%20States&utm_content=22066&utm_campaign=profiles_installer'

            self.browser.get(url)
            data_list = [url]
            try:
                name = self.browser.find_element_by_css_selector(
                    'body > div.container > div > div > div.enf-company-profile-info.clearfix > div.enf-company-profile-info-main.pull-left > h1').text
            except:
                name = ''
            try:
                address = self.browser.find_element_by_css_selector(
                    'body > div.container > div > div > div.enf-company-profile-info.clearfix > div.enf-company-profile-info-main.pull-left > div > table:nth-child(1) > tbody > tr > td:nth-child(2)').text
            except:
                address = ''

            try:
                phone = self.browser.find_element_by_css_selector(
                    'body > div.container > div > div > div.enf-company-profile-info.clearfix > div.enf-company-profile-info-main.pull-left > div > table:nth-child(2) > tbody > tr > td.ar\:number-direction > a').text
            except:
                phone = ''

            try:
                website = self.browser.find_element_by_css_selector(
                    'body > div.container > div > div > div.enf-company-profile-info.clearfix > div.enf-company-profile-info-main.pull-left > div > table:nth-child(3) > tbody > tr > td:nth-child(2) > a').get_attribute(
                    'href')
            except:
                website = ''

            try:
                country = self.browser.find_element_by_css_selector(
                    'body > div.container > div > div > div.enf-company-profile-info.clearfix > div.enf-company-profile-info-main.pull-left > div > table:nth-child(4) > tbody > tr > td:nth-child(2)').text
            except:
                country = ''

            on_off_grid = ''
            installation_size = ''
            other_services = ''
            operating_area = ''
            panel_suppliers = ''
            last_update = ''
            try:
                table_one = self.browser.find_elements_by_css_selector('#installer > div > div > div')
                for row in table_one:
                    title = row.find_element_by_css_selector('div.col-xs-2.enf-section-body-title').text
                    data = row.find_element_by_css_selector('div.col-xs-10.enf-section-body-content.blue').text
                    if 'On-Grid / Off-Grid' in title:
                        on_off_grid = data
                        print('On-Grid / Off-Grid : ', on_off_grid)
                    if 'Installation size' in title:
                        installation_size = data
                        print('installation_size : ', installation_size)
                    if 'Other Services' in title:
                        other_services = data
                        print('other_services : ', other_services)
                    if 'Operating Area' in title:
                        operating_area = data
                        print('operating_area : ', operating_area)
                    if 'Panel Suppliers' in title:
                        panel_suppliers = data
                        print('panel_suppliers : ', panel_suppliers)
                last_update = self.browser.find_element_by_css_selector(
                    'body > div.container > div > div > div.enf-section.enf-section-update-area > div > div > div.col-xs-10.enf-section-body-content').text

            except:
                pass

            sales_contracts = ''
            expansions = ''
            financial_news = ''
            other_news = ''
            try:
                table_two = self.browser.find_elements_by_css_selector(
                    'body > div.container > div > div > div.enf-section.enf-section-news > div > div')
                for row in table_two:
                    title = row.find_element_by_css_selector('div:nth-child(1)').text
                    row.find_element_by_css_selector('div:nth-child(1)').click()
                    data = row.find_element_by_css_selector('div:nth-child(2)').text
                    sleep(2)
                    if 'Sales Contracts' in title:
                        try:
                            sales_contracts = data
                        except:
                            pass
                    if 'Expansions' in title:
                        try:
                            expansions = data
                        except:
                            pass
                    if 'Financial News' in title:
                        try:
                            financial_news = data
                        except:
                            pass
                    if 'Other News' in title:
                        try:
                            other_news = data
                        except:
                            pass
            except:
                pass

            data_list.extend(
                [name, address, phone, website, country, on_off_grid, installation_size, other_services, operating_area,
                 panel_suppliers, last_update, sales_contracts, expansions, financial_news, other_news])
            self.write_to_table(self.index, 1, data_list)
            print(self.index - 1, data_list)
            self.index += 1
        except Exception as ex:
            print(ex)


instance = SolarSystemScript()
instance.main()